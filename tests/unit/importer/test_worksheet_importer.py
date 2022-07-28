from unittest import TestCase
from unittest.mock import MagicMock

from openpyxl import Workbook

from hca_ingest.importer.conversion.metadata_entity import MetadataEntity
from hca_ingest.importer.importer import WorksheetImporter
from hca_ingest.importer.spreadsheet.ingest_worksheet import IngestWorksheet


class WorksheetImporterTest(TestCase):

    def test_do_import(self):
        # given:
        row_template = MagicMock('row_template')
        no_errors = []

        # and:
        john_doe = MetadataEntity(object_id='profile_1')
        emma_jackson = MetadataEntity(object_id='profile_2')
        row_template.do_import = MagicMock('import_row', side_effect=[(john_doe, no_errors), (emma_jackson, no_errors)])

        # and:
        mock_template_manager = MagicMock('template_manager')
        mock_template_manager.create_row_template = MagicMock(return_value=row_template)
        mock_template_manager.get_header_row = MagicMock(return_value=['header1', 'header2'])
        mock_template_manager.get_concrete_type = MagicMock(return_value='concrete_entity')

        # and:
        workbook = Workbook()
        worksheet = workbook.create_sheet('user_profile')
        worksheet['A4'] = 'header'
        worksheet['A6'] = 'john'
        worksheet['A7'] = 'emma'

        # when:
        worksheet_importer = WorksheetImporter(mock_template_manager)
        profiles, errors = worksheet_importer.do_import(IngestWorksheet(worksheet))

        # then:
        self.assertEqual(2, len(profiles))
        self.assertIn(john_doe, profiles)
        self.assertIn(emma_jackson, profiles)
        self.assertEqual(errors, [])

        # and: domain and concrete type should be set
        pass

    def test_do_import_no_id_metadata(self):
        # given:
        row_template = MagicMock('row_template')
        no_errors = []
        # and:
        paper_metadata = MetadataEntity(content={'product_name': 'paper'},
                                        links={'delivery': ['123', '456']})
        pen_metadata = MetadataEntity(content={'product_name': 'pen'},
                                      links={'delivery': ['789']})
        row_template.do_import = MagicMock(side_effect=([(paper_metadata, no_errors), (pen_metadata, no_errors)]))

        # and:
        mock_template_manager = MagicMock('template_manager')
        mock_template_manager.create_row_template = MagicMock(return_value=row_template)
        mock_template_manager.get_header_row = MagicMock(return_value=['header1', 'header2'])
        mock_template_manager.get_concrete_type = MagicMock(return_value='concrete_entity')

        # and:
        workbook = Workbook()
        worksheet = workbook.create_sheet('product')
        worksheet['A6'] = 'paper'
        worksheet['A7'] = 'pen'

        # when:
        worksheet_importer = WorksheetImporter(mock_template_manager)
        results, errors = worksheet_importer.do_import(IngestWorksheet(worksheet))

        # then:
        self.assertEqual(2, len(results))
        self.assertIn(paper_metadata, results)
        self.assertIn(pen_metadata, results)
        self.assertEqual(errors, [])

        # and: object id should be assigned
        paper_id = paper_metadata.object_id
        self.assertIsNotNone(paper_id)
        pen_id = pen_metadata.object_id
        self.assertIsNotNone(pen_id)
        self.assertNotEqual(paper_id, pen_id)
