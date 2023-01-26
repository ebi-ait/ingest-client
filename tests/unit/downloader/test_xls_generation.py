import pytest

from assertpy import assert_that
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from hca_ingest.downloader.downloader import (XlsDownloader, HEADER_ROW_NO, BORDER_ROW_NO, TITLE_FONT, TITLE_ALIGNMENT,
                                              TITLE_FILL, DESCRIPTION_FONT, DESCRIPTION_ALIGNMENT, BORDER_VALUE, HEADER_PROTECTION)
from hca_ingest.importer.spreadsheet.ingest_workbook import SCHEMAS_WORKSHEET

from .conftest import get_json_file


@pytest.fixture
def downloader():
    return XlsDownloader()


@pytest.fixture
def project_json(blank_header):
    return {
        'Project' : {
            "headers": {
                "project.uuid": blank_header,
                "project.project_core.project_short_name": blank_header,
                "project.project_core.project_title": blank_header,
                "project.project_core.project_description": blank_header,
                "project.insdc_project_accessions": blank_header,
                "project.geo_series_accessions": blank_header,
                "project.insdc_study_accessions": blank_header
            },
            "values": [
                {
                    "project.uuid": "3e329187-a9c4-48ec-90e3-cc45f7c2311c",
                    "project.project_core.project_short_name": "kriegsteinBrainOrganoids",
                    "project.project_core.project_title": "Establishing Cerebral Organoids as Models of Human-Specific Brain Evolution",
                    "project.project_core.project_description": "Direct comparisons of human and non-human primate brain tissue have the potential to reveal molecular pathways underlying remarkable specializations of the human brain. However, chimpanzee tissue is largely inaccessible during neocortical neurogenesis when differences in brain size first appear. To identify human-specific features of cortical development, we leveraged recent innovations that permit generating pluripotent stem cell-derived cerebral organoids from chimpanzee. First, we systematically evaluated the fidelity of organoid models to primary human and macaque cortex, finding organoid models preserve gene regulatory networks related to cell types and developmental processes but exhibit increased metabolic stress. Second, we identified 261 genes differentially expressed in human compared to chimpanzee organoids and macaque cortex. Many of these genes overlap with human-specific segmental duplications and a subset suggest increased PI3K/AKT/mTOR activation in human outer radial glia. Together, our findings establish a platform for systematic analysis of molecular changes contributing to human brain development and evolution. Overall design: Single cell mRNA sequencing of iPS-derived neural and glial progenitor cells using the Fluidigm C1 system This series includes re-analysis of publicly available data in accessions: phs000989.v3, GSE99951, GSE86207, GSE75140. Sample metadata and accession IDs for the re-analyzed samples are included in the file \"GSE124299_metadata_on_processed_samples.xlsx\" available on the foot of this record. The following samples have no raw data due to data loss: GSM3569728, GSM3569738, GSM3571601, GSM3571606, GSM3571615, GSM3571621, GSM3571625, and GSM3571631",
                    "project.insdc_project_accessions": "SRP180337",
                    "project.geo_series_accessions": "GSE124299",
                    "project.insdc_study_accessions": "PRJNA515930"
                }
            ]
        },
        "Schemas": ["dummy-schema-url"]
    }


@pytest.fixture
def contributors_json(blank_header):
    return {
        'Project - Contributors': {
            'headers': {
                "project.contributors.name": blank_header,
                "project.contributors.email": blank_header,
                "project.contributors.institution": blank_header,
                "project.contributors.laboratory": blank_header,
                "project.contributors.address": blank_header,
                "project.contributors.country": blank_header,
                "project.contributors.corresponding_contributor": blank_header,
                "project.contributors.project_role.text": blank_header,
                "project.contributors.project_role.ontology": blank_header,
                "project.contributors.project_role.ontology_label": blank_header,
                "project.contributors.orcid_id": blank_header
            },
            'values': [
                {
                    'project.contributors.name': 'Karel Gott',
                    'project.contributors.email': 'karel@gott.com',
                    'project.contributors.institution': 'University of Prague',
                    'project.contributors.laboratory': 'Department of Neurology',
                    'project.contributors.address': '123 Test Street',
                    'project.contributors.country': 'UK',
                    'project.contributors.corresponding_contributor': 'yes',
                    'project.contributors.project_role.text': 'experimental scientist',
                    'project.contributors.project_role.ontology': 'EFO:0009741',
                    'project.contributors.project_role.ontology_label': 'experimental scientist',
                    'project.contributors.orcid_id': 'https://orcid.org/1234-5678-9012-3456'
                },
                {
                    'project.contributors.name': 'Lady Carneval',
                    'project.contributors.email': 'ladyc@gott.com',
                    'project.contributors.institution': 'University of Prague',
                    'project.contributors.laboratory': 'Department of Neurology',
                    'project.contributors.address': '123 Test Street',
                    'project.contributors.country': 'UK',
                    'project.contributors.corresponding_contributor': 'yes',
                    'project.contributors.project_role.text': 'experimental scientist',
                    'project.contributors.project_role.ontology': 'EFO:0009741',
                    'project.contributors.project_role.ontology_label': 'experimental scientist',
                    'project.contributors.orcid_id': 'https://orcid.org/9999-9999-9999-9999'
                },
                {
                    'project.contributors.name': 'Baby Shark',
                    'project.contributors.email': 'sharkb@gott.com',
                    'project.contributors.institution': 'University of Cambridge',
                    'project.contributors.laboratory': 'Department of Neurology',
                    'project.contributors.address': '123 Ocean Drive',
                    'project.contributors.country': 'USA',
                    'project.contributors.corresponding_contributor': 'no',
                    'project.contributors.project_role.text': 'data wrangler',
                    'project.contributors.project_role.ontology': 'EFO:0009737',
                    'project.contributors.project_role.ontology_label': 'data curator',
                    'project.contributors.orcid_id': 'https://orcid.org/0000-0000-0000-0001'
                }
            ]
        },
        "Schemas": ["dummy-schema-url"]
    }


@pytest.fixture
def multi_line_project(script_dir):
    return get_json_file(script_dir+ '/project-list-flattened.json')


@pytest.fixture(params=[
    pytest.lazy_fixture('project_json'),
    pytest.lazy_fixture('contributors_json'),
    pytest.lazy_fixture('multi_line_project'),
])
def without_schema_headers(request):
    return request.param


@pytest.fixture
def with_schema_headers(script_dir):
    return get_json_file(script_dir + '/small-project-flattened-with-schema.json')


@pytest.fixture(params=[
    pytest.lazy_fixture('without_schema_headers'),
    pytest.lazy_fixture('with_schema_headers'),
])
def input_json(request):
    return request.param


def test_xls_downloader(downloader, input_json):
    # when
    workbook: Workbook = downloader.create_workbook(input_json)
    # expect
    assert_workbook(workbook, input_json)


def test_given_input_raises_error_when_no_schemas_worksheet(downloader):
    with pytest.raises(ValueError) as value_error:
        downloader.create_workbook({})
        assert_that(str(value_error.value)).is_equal_to("The schema urls are missing")


def assert_workbook(workbook: Workbook, input_json: dict):
    assert_schema(workbook, input_json)
    sheet_names = [sheet_name for sheet_name in input_json.keys() if sheet_name != 'Schemas']
    for sheet_name in sheet_names:
        assert_sheet(workbook, sheet_name, input_json)


def assert_sheet(workbook: Workbook, sheet_title: str, input_json :dict):
    sheet = workbook[sheet_title]
    input_sheet = input_json[sheet_title]
    input_headers = input_sheet['headers']
    input_values = input_sheet['values']

    assert_headers(sheet, input_headers)
    assert_values(sheet, input_values)


def assert_headers(worksheet: Worksheet, input_headers: dict):
    border_cell = worksheet[f'A{BORDER_ROW_NO}']
    assert_that(border_cell).has_value(BORDER_VALUE).has_font(TITLE_FONT).has_fill(TITLE_FILL)
    for user_friendly, description, guide, key in worksheet.iter_cols(max_row=HEADER_ROW_NO):
        assert_that(input_headers).contains_key(key.value)
        input_header = input_headers[key.value]
        assert_that(key).has_protection(HEADER_PROTECTION)
        assert_that(user_friendly).has_font(TITLE_FONT).has_fill(TITLE_FILL).has_alignment(TITLE_ALIGNMENT)
        if input_header['user_friendly']:
            assert_that(user_friendly.value).starts_with(input_header['user_friendly'])
        if input_header['required']:
            assert_that(user_friendly.value).ends_with(' (Required)')
        assert_that(description).has_font(DESCRIPTION_FONT).has_alignment(DESCRIPTION_ALIGNMENT).has_value(input_header['description'])
        assert_that(guide).has_font(DESCRIPTION_FONT).has_alignment(DESCRIPTION_ALIGNMENT)
        if input_header['guidelines']:
            assert_that(guide.value).starts_with(input_header['guidelines'])
        if input_header['example']:
            assert_that(guide.value).ends_with(f'For example: {input_header["example"]}')


def assert_values(worksheet: Worksheet, input_values: list[dict]):
    for index, row in enumerate(worksheet.iter_rows(min_row=BORDER_ROW_NO + 1)):
        input_value = input_values[index]
        cell: Cell
        for cell in row:
            column_name = worksheet[f'{cell.column_letter}{HEADER_ROW_NO}'].value
            if column_name in input_value:
                assert_that(cell.value).is_equal_to(input_value[column_name])
            else:
                assert_that(cell.value).is_none()


def assert_schema(workbook, flat_json):
    assert_that(workbook.sheetnames).contains(SCHEMAS_WORKSHEET)
    expected_schemas = flat_json.get(SCHEMAS_WORKSHEET)
    sheet: Worksheet = workbook[SCHEMAS_WORKSHEET]
    rows_iter = sheet.iter_rows(min_col=1, min_row=2, max_col=1, max_row=sheet.max_row)
    schemas = []
    for row in rows_iter:
        for cell in row:
            schemas.append(cell.value)
    assert_that(schemas).is_equal_to(expected_schemas)
