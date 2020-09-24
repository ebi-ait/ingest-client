import string
from unittest import TestCase

from mock import Mock
from openpyxl import Workbook

from ingest.importer.spreadsheet.ingest_workbook import IngestWorkbook
from ingest.importer.submission import Entity
from tests.importer.utils.test_utils import create_ingest_workbook, create_test_workbook,\
    create_mocked_workbook_importer


class IngestWorkbookTest(TestCase):

    def test_get_schemas(self):
        # given:
        workbook = Workbook()

        # and:
        schemas_sheet = workbook.create_sheet('Schemas')
        schemas_sheet['A1'] = 'schema'

        # and:
        base_url = 'https://schema.humancellatlas.org'
        expected_schemas = [
            f'{base_url}/type/biomaterial/cell_suspension',
            f'{base_url}type/biomaterial/organ_from_donor',
            f'{base_url}/type/process/library_preparation'
        ]
        header_offset = 2
        for index, schema in enumerate(expected_schemas):
            schemas_sheet[f'A{index + header_offset}'] = schema

        # and:
        ingest_workbook = IngestWorkbook(workbook)

        # when:
        actual_schemas = ingest_workbook.get_schemas()

        # then:
        self.assertEqual(expected_schemas, actual_schemas)

    def test_importable_worksheets(self):
        # given:
        importable_names = ['Organ From Donor', 'Cell Suspension', 'Project']
        workbook = create_test_workbook(*importable_names)
        workbook.create_sheet('Schemas')

        # and:
        ingest_workbook = IngestWorkbook(workbook)

        # when:
        actual_worksheets = ingest_workbook.importable_worksheets()

        # then:
        actual_titles = [ingest_worksheet.title for ingest_worksheet in actual_worksheets]
        self.assertEqual(importable_names, actual_titles)

    def test_add_entity_uuids(self):
        # given:
        entities = [Entity('type', 'AA', 'content', concrete_type='a', ingest_json={'uuid': {'uuid': 'A-1-uuid'}},
                           spreadsheet_location={'row_index': 5, 'worksheet_title': 'A'}),
                    Entity('type', 'AA', 'content', concrete_type='a', ingest_json={'uuid': {'uuid': 'A-2-uuid'}},
                           spreadsheet_location={'row_index': 6, 'worksheet_title': 'A'}),
                    Entity('type', 'AA', 'content', concrete_type='a', ingest_json={'uuid': {'uuid': 'A-3-uuid'}},
                           spreadsheet_location={'row_index': 7, 'worksheet_title': 'A'}),
                    Entity('type', 'BB', 'content', concrete_type='b', ingest_json={'uuid': {'uuid': 'B-1-uuid'}},
                           spreadsheet_location={'row_index': 5, 'worksheet_title': 'B'}),
                    Entity('type', 'BB', 'content', concrete_type='b', ingest_json={'uuid': {'uuid': 'B-2-uuid'}},
                           spreadsheet_location={'row_index': 6, 'worksheet_title': 'B'}),
                    Entity('type', 'BB', 'content', concrete_type='b', ingest_json={'uuid': {'uuid': 'B-3-uuid'}},
                           spreadsheet_location={'row_index': 7, 'worksheet_title': 'B'}),
                    Entity('type', 'CC', 'content', concrete_type='c', ingest_json={'uuid': {'uuid': 'C-1-uuid'}},
                           spreadsheet_location={'row_index': 5, 'worksheet_title': 'C'}),
                    Entity('type', 'CC', 'content', concrete_type='c', ingest_json={'uuid': {'uuid': 'C-2-uuid'}},
                           spreadsheet_location={'row_index': 6, 'worksheet_title': 'C'}),
                    Entity('type', 'CC', 'content', concrete_type='c', ingest_json={'uuid': {'uuid': 'C-3-uuid'}},
                           spreadsheet_location={'row_index': 7, 'worksheet_title': 'C'}),
                    Entity('type', 'XX', 'content', concrete_type='x', ingest_json={'uuid': {'uuid': 'X-3-uuid'}})]

        mock_submission = Mock('submission')
        mock_submission.get_entities = Mock(return_value=entities)

        sheets = ['A', 'B', 'C', 'X']
        workbook = create_test_workbook(*sheets)
        wb = IngestWorkbook(workbook)

        # when
        wb.add_entity_uuids(mock_submission)

        # then
        self.assertEqual(wb.workbook['A'].cell(4, 1).value, 'a.uuid')
        self.assertEqual(wb.workbook['A'].cell(5, 1).value, 'A-1-uuid')
        self.assertEqual(wb.workbook['A'].cell(6, 1).value, 'A-2-uuid')
        self.assertEqual(wb.workbook['A'].cell(7, 1).value, 'A-3-uuid')

        self.assertEqual(wb.workbook['B'].cell(4, 1).value, 'b.uuid')
        self.assertEqual(wb.workbook['B'].cell(5, 1).value, 'B-1-uuid')
        self.assertEqual(wb.workbook['B'].cell(6, 1).value, 'B-2-uuid')
        self.assertEqual(wb.workbook['B'].cell(7, 1).value, 'B-3-uuid')

        self.assertEqual(wb.workbook['C'].cell(4, 1).value, 'c.uuid')
        self.assertEqual(wb.workbook['C'].cell(5, 1).value, 'C-1-uuid')
        self.assertEqual(wb.workbook['C'].cell(6, 1).value, 'C-2-uuid')
        self.assertEqual(wb.workbook['C'].cell(7, 1).value, 'C-3-uuid')

        self.assertFalse(wb.workbook['X'].cell(5, 1).value, None)

    def test_add_schemas_worksheet(self):
        # given
        sheets = ['A', 'B', 'C', 'X']
        workbook = create_test_workbook(*sheets)
        wb = IngestWorkbook(workbook)

        # when
        wb.add_schemas_worksheet(['schema1', 'schema2', 'schema3'])

        # then
        self.assertEqual(wb.workbook['Schemas'].cell(1, 1).value, 'Schemas')
        self.assertEqual(wb.workbook['Schemas'].cell(2, 1).value, 'schema1')
        self.assertEqual(wb.workbook['Schemas'].cell(3, 1).value, 'schema2')
        self.assertEqual(wb.workbook['Schemas'].cell(4, 1).value, 'schema3')

    def test_add_schemas_worksheet_existing_schemas(self):
        # given
        sheets = ['A', 'B', 'C']
        workbook = create_test_workbook(*sheets)
        wb = IngestWorkbook(workbook)

        wb.workbook.create_sheet('Schemas')
        wb.workbook['Schemas'].cell(1, 1).value = 'X'
        wb.workbook['Schemas'].cell(2, 1).value = 'x'
        wb.workbook['Schemas'].cell(3, 1).value = 'y'
        wb.workbook['Schemas'].cell(4, 1).value = 'z'

        # when
        wb.add_schemas_worksheet(['schema1', 'schema2', 'schema3'])

        # then
        self.assertEqual(wb.workbook['Schemas'].cell(1, 1).value, 'X')
        self.assertEqual(wb.workbook['Schemas'].cell(2, 1).value, 'x')
        self.assertEqual(wb.workbook['Schemas'].cell(3, 1).value, 'y')
        self.assertEqual(wb.workbook['Schemas'].cell(4, 1).value, 'z')

    def test_when_proceeding_update_and_all_entities_has_uuids_then_return_no_errors(self):
        # given:
        is_update = True
        person = 'person'
        workbook_importer = create_mocked_workbook_importer([person])

        # and:
        ingest_workbook = create_ingest_workbook([person], ['uuid', 'name'])

        # then:
        self.assertTrue(
            len(workbook_importer.validate_worksheets(is_update, ingest_workbook.importable_worksheets())) == 0
        )

    def test_when_proceeding_update_and_some_entities_has_no_uuids_then_return_errors(self):
        # given:
        is_update = True
        person = 'person'
        workbook_importer = create_mocked_workbook_importer([person])

        # and:
        ingest_workbook = create_ingest_workbook([person], ['name', 'address'])

        # and:
        expected_error = {
            'location': f'sheet={person}',
            'type': 'MissingEntityUUIDFound',
            'detail': f'The {person} entities in the spreadsheet should have UUIDs.'}
        # then:
        worksheet_errors = workbook_importer.validate_worksheets(is_update, ingest_workbook.importable_worksheets())
        self.assertTrue(len(worksheet_errors) > 0)
        self.assertIn(expected_error, worksheet_errors)

    def test_when_proceeding_create_and_some_entities_has_got_uuids_then_return_errors(self):
        # given:
        is_update = False
        person = 'person'
        workbook_importer = create_mocked_workbook_importer([person])

        # and:
        ingest_workbook = create_ingest_workbook([person], ['uuid', 'address'])

        # and:
        expected_error = {
            'location': f'sheet={person}',
            'type': 'UnexpectedEntityUUIDFound',
            'detail': f'The {person} entities in the spreadsheet shouldn’t have UUIDs.'}
        # then:
        worksheet_errors = workbook_importer.validate_worksheets(is_update, ingest_workbook.importable_worksheets())
        self.assertTrue(len(worksheet_errors) > 0)
        self.assertIn(expected_error, worksheet_errors)

    def test_when_proceeding_create_and_no_entities_has_uuids_then_return_no_errors(self):
        # given:
        is_update = False
        person = 'person'
        workbook_importer = create_mocked_workbook_importer([person])

        # and:
        ingest_workbook = create_ingest_workbook([person], ['name', 'address'])

        # then:
        worksheet_errors = workbook_importer.validate_worksheets(is_update, ingest_workbook.importable_worksheets())
        self.assertTrue(len(worksheet_errors) == 0)
