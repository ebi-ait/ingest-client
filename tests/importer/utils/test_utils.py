from openpyxl import Workbook
from mock import Mock, MagicMock
from ingest.importer.importer import WorkbookImporter
from ingest.importer.spreadsheet.ingest_workbook import IngestWorkbook


def create_test_workbook(*worksheet_titles, include_default_sheet=False):
    workbook = Workbook()
    for title in worksheet_titles:
        workbook.create_sheet(title)

    check_default_sheet_inclusion(include_default_sheet, workbook)

    return workbook


def create_workbook_importer(entity_names):
    template_manager = MagicMock()
    concrete_type_map = {}
    for entity_name in entity_names:
        concrete_type_map[entity_name] = entity_name
    template_manager.get_concrete_type = lambda key: concrete_type_map.get(key)

    return WorkbookImporter(template_manager)


def create_ingest_workbook(sheet_names, column_names, include_default_sheet=False):
    header_row_idx = 4

    workbook = Workbook()
    for sheet_name in sheet_names:
        worksheet = workbook.create_sheet(sheet_name)
        worksheet[f'A{header_row_idx}'] = f'{sheet_name}.{column_names[0]}'
        worksheet[f'B{header_row_idx}'] = f'{sheet_name}.{column_names[1]}'

    check_default_sheet_inclusion(include_default_sheet, workbook)

    return IngestWorkbook(workbook)


def check_default_sheet_inclusion(include_default_sheet, workbook):
    if not include_default_sheet:
        default_sheet = workbook['Sheet']
        workbook.remove(default_sheet)
