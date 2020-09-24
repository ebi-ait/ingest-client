import os

from ingest.template.schema_template import SchemaTemplate

__author__ = "jupp"
__license__ = "Apache 2.0"
__date__ = "25/05/2018"

from unittest import TestCase
from ingest.template.vanilla_spreadsheet_builder import VanillaSpreadsheetBuilder
import unittest
from openpyxl import Workbook
from openpyxl import load_workbook as Reader



class TestSchemaTemplate(TestCase):

    def setUp(self):
        self.longMessage = True
        self.projectUri = "https://schema.humancellatlas.org/type/project/5.1.0/project"
        self.donorUri = "https://schema.humancellatlas.org/type/biomaterial/5.1.0/donor_organism"

    def test_no_schemas(self):
        data = {
            "id": self.donorUri,
            "properties": {
                "foo_bar": {
                    "user_friendly": "Foo bar",
                    "description": "this is a foo bar",
                    "example": "e.g. foo"
                }
            }
        }

        file = "foo.xlsx"
        spreadsheet_builder = VanillaSpreadsheetBuilder(file)
        template = SchemaTemplate(json_schema_docs=[data], property_migrations=[])
        spreadsheet_builder.build(template)
        spreadsheet_builder.save_spreadsheet()

        reader = Reader("foo.xlsx")
        sheet = reader["Donor organism"]

        self.assertEqual("this is a foo bar", sheet.cell(row=2, column=1).value)
        self.assertEqual("FOO BAR", sheet.cell(row=1, column=1).value)
        self.assertEqual("For example: e.g. foo", sheet.cell(row=3, column=1).value.strip())
        self.assertEqual("donor_organism.foo_bar", sheet.cell(row=4, column=1).value)
        # clean up
        os.remove(file)

    def test_user_friendly(self):
        user_friendly_dict = {
            "donor_organism.human_specific.body_mass_index": "Body mass index",
            "specimen_from_organism.purchased_specimen.manufacturer": "Purchased specimen - Manufacturer",
            "donor_organism.organism_age_unit.text": "Age unit",
            "donor_organism.organism_age_unit.ontology": "Age unit ontology ID",
            "library_preparation_protocol.cell_barcode.barcode_length": "Cell barcode - Barcode length",
            "project.contributors.project_role.ontology_label": "Contributor role ontology label",
            "donor_organism.human_specific.ethnicity.text": "Ethnicity",
            "collection_protocol.reagents.retail_name": "Retail name",
            "imaging_protocol.probe.probe_reagents.catalog_number": "Catalog number",
            "donor_organism.genus_species.text": "Genus species",
            "donor_organism.genus_species.ontology": "Genus species ontology ID",
            "donor_organism.genus_species.ontology_label": "Genus species ontology label",
            "cell_suspension.cell_morphology.cell_size_unit.text": "Cell size unit",
            "specimen_from_organism.preservation_storage.storage_time_unit.text": "Storage time unit",
            "cell_suspension.timecourse.unit.text": "Timecourse unit"
        }

        file = "uf_test.xlsx"
        template = SchemaTemplate()
        builder = VanillaSpreadsheetBuilder(file)
        for key in user_friendly_dict.keys():
            uf = builder.get_user_friendly_column_name(template, key)
            print("from method: " + uf)
            print("expected: " + user_friendly_dict[key])
            self.assertEqual(user_friendly_dict[key], uf)


    # TODO improve this test, at the moment just tests that the spreadsheet can be built, nothing about the
    #  contents.
    def test_vanilla_spreadsheet(self):
        file = "uf_test.xlsx"
        template = SchemaTemplate()
        builder = VanillaSpreadsheetBuilder(file)
        builder.generate_spreadsheet(schema_template=template)
        builder.save_spreadsheet()
        reader = Reader("uf_test.xlsx")
        print(type(reader))
        self.assertIsInstance(reader, Workbook)
        # cleanup
        os.remove(file)

    def test_correct_description_used(self):
        file = "uf_test.xlsx"
        template = SchemaTemplate()
        builder = VanillaSpreadsheetBuilder(file)
        test_field = "enrichment_protocol.method.text"
        returned_description = builder.get_value_for_column(template=template, column_name=test_field,
                                                            property="description")
        print("returned description: " + returned_description)
        expected_description = "The method by which enrichment was achieved."
        returned_example_text = builder.get_value_for_column(template=template, column_name=test_field,
                                                             property="example")
        print("returned_example_text: " + returned_example_text)
        expected_example_text = "enzymatic dissociation; blood draw"
        self.assertEqual(expected_description, returned_description)
        self.assertEqual(expected_example_text, returned_example_text)

    # TODO fixme
    @unittest.skip
    def test_with_tabs_template(self):
        pass
        # spreadsheet_builder.generate_spreadsheet("human_10x.xlsx", tabs_template="tabs_human_10x.yaml",
        #                                          schema_urls=schemas)

    # TODO fixme
    @unittest.skip
    def test_add_columns(self):
        # spreadsheet_builder.generate_spreadsheet("generic.xlsx", schema_urls=schemas)
        pass

    # TODO fixme
    @unittest.skip
    def test_add_sheets(self):
        # spreadsheet_builder.generate_spreadsheet("generic.xlsx", schema_urls=schemas)
        pass
