from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from hca_ingest.importer.spreadsheet.ingest_workbook import SCHEMAS_WORKSHEET
from hca_ingest.importer.spreadsheet.ingest_worksheet import START_DATA_ROW

USER_FRIENDLY_ROW_NO = 1
DESCRIPTION_ROW_NO = 2
GUIDELINES_ROW_NO = 3
HEADER_ROW_NO = 4
BOARDER_ROW_NO = 5

class XlsDownloader:
    @staticmethod
    def create_workbook(input_json: dict) -> Workbook:
        workbook = Workbook()
        workbook.remove(workbook.active)

        for ws_title, ws_elements in input_json.items():
            if ws_title == 'Project':
                worksheet: Worksheet = workbook.create_sheet(title=ws_title, index=0)
            elif ws_title == SCHEMAS_WORKSHEET:
                continue
            else:
                worksheet: Worksheet = workbook.create_sheet(title=ws_title)
            XlsDownloader.add_worksheet_content(worksheet, ws_elements)

        XlsDownloader.generate_schemas_worksheet(input_json, workbook)
        return workbook

    @staticmethod
    def generate_schemas_worksheet(input_json: dict, workbook: Workbook):
        schemas = input_json.get(SCHEMAS_WORKSHEET)
        if not schemas:
            raise ValueError('The schema urls are missing')
        schemas_worksheet = workbook.create_sheet(SCHEMAS_WORKSHEET)
        schemas_worksheet.cell(row=1, column=1, value=SCHEMAS_WORKSHEET)
        for row_num, schema in enumerate(schemas, start=2):
            schemas_worksheet.cell(row=row_num, column=1, value=schema)

    @staticmethod
    def add_worksheet_content(worksheet: Worksheet, ws_elements: dict):
        headers = ws_elements.get('headers', {})
        XlsDownloader.__add_header_rows(worksheet, headers)
        all_values = ws_elements.get('values', [])

        for row_number, row_values in enumerate(all_values, start=START_DATA_ROW):
            XlsDownloader.__add_row_content(worksheet, headers, row_number, row_values)

    @staticmethod
    def __add_header_rows(worksheet, headers: dict):
        worksheet.cell(row=BOARDER_ROW_NO, column=1, value='FILL OUT INFORMATION BELOW THIS ROW')
        for col, header in enumerate(headers.keys(), start=1):
            XlsDownloader.__add_column_header(worksheet, col, header, headers.get(header, {}))

    @staticmethod
    def __add_column_header(worksheet, column_number: int, column_key: str, header_info: dict):
        user_friendly = header_info.get('user_friendly', '')
        if header_info.get('required', False):
            user_friendly = f'{user_friendly} (Required)'
        description = header_info.get('description', '')
        example = header_info.get('example', '')
        guidelines = header_info.get('guidelines', '')
        if example:
            guidelines = f'{guidelines} For example: {example}'
        worksheet.cell(row=USER_FRIENDLY_ROW_NO, column=column_number, value=user_friendly)
        worksheet.cell(row=DESCRIPTION_ROW_NO, column=column_number, value=description)
        worksheet.cell(row=GUIDELINES_ROW_NO, column=column_number, value=guidelines)
        worksheet.cell(row=HEADER_ROW_NO, column=column_number, value=column_key)

    @staticmethod
    def __add_row_content(worksheet, headers: dict, row_number: int, values: dict):
        for index, header in enumerate(headers.keys(), start=1):
            if header in values:
                worksheet.cell(row=row_number, column=index, value=values[header])
